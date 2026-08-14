"""Parser Excel SSW 081 — CTRCs disponíveis para entrega (Sem saída).

Colunas fixas (Excel):
  B=CTRC · J=CLIENTE · E=DATA PREVISTA · M=CIDADE · N=UF · X=ULTIMA OCORRENCIA
Frete/peso: descobertos pelo cabeçalho quando existirem.
"""
from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from config import BASE_DIR, ensure_dirs

CACHE_DIR = BASE_DIR / "data" / "cache"
ITENS_081_CSV = CACHE_DIR / "itens_081.csv"
RESUMO_081_CSV = CACHE_DIR / "resumo_081.csv"
TOP_CTE_081_CSV = CACHE_DIR / "top_cte_081.csv"
TOP_CLIENTE_081_CSV = CACHE_DIR / "top_cliente_081.csv"
LAST_081_JSON = CACHE_DIR / "last_run_081.json"

COL_CTRC = "B"
COL_CLIENTE = "J"
COL_PREVISTA = "E"
COL_CIDADE = "M"
COL_UF = "N"
COL_OCORRENCIA = "X"

ITEM_FIELDS = [
    "ctrc",
    "cliente",
    "data_prevista",
    "dias_atraso",
    "cidade",
    "uf",
    "cidade_uf",
    "ultima_ocorrencia",
    "frete",
    "peso",
]
RESUMO_FIELDS = [
    "atualizado",
    "periodo",
    "qtd",
    "frete",
    "frete_fmt",
    "peso",
    "peso_fmt",
]
TOP_CTE_FIELDS = ["ctrc", "cliente", "dias_atraso", "cidade_uf", "ultima_ocorrencia"]
TOP_CLIENTE_FIELDS = ["cliente", "dias_max", "qtd", "frete", "peso"]

TOP_N = 10


def _col_letter_to_idx(letters: str) -> int:
    n = 0
    for ch in str(letters or "").strip().upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - ord("A") + 1)
    return max(0, n - 1)


def _cell_by_letter(row: tuple[Any, ...] | list[Any], letters: str) -> Any:
    idx = _col_letter_to_idx(letters)
    if idx >= len(row):
        return ""
    return row[idx]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            pass
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    try:
        n = float(text or 0)
        return -n if neg else n
    except Exception:
        return 0.0


def _fmt_int(n: float | int) -> str:
    return f"{int(round(float(n) or 0)):,}".replace(",", ".")


def _fmt_dec(n: float, places: int = 2) -> str:
    s = f"{float(n or 0):,.{places}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_money(n: float) -> str:
    return f"R$ {_fmt_dec(n, 2)}"


def _norm_header(h: Any) -> str:
    t = _clean(h).lower()
    for a, b in (
        ("ç", "c"),
        ("ã", "a"),
        ("á", "a"),
        ("à", "a"),
        ("â", "a"),
        ("é", "e"),
        ("ê", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ô", "o"),
        ("ú", "u"),
    ):
        t = t.replace(a, b)
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def _find_metric_cols(headers: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        n = _norm_header(h)
        if not n:
            continue
        if "frete" in n or n in {"vlr frete", "valor frete", "vl frete"}:
            out.setdefault("frete", i)
        elif n == "peso" or n.startswith("peso ") or "peso real" in n or "peso taxado" in n:
            out.setdefault("peso", i)
    return out


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, "date") and callable(getattr(value, "date")):
        try:
            d = value.date()
            if isinstance(d, date):
                return d
        except Exception:
            return None
    text = _clean(value)
    if not text:
        return None
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if m:
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return date(yy, mm, dd)
        except Exception:
            return None
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        try:
            return date(int(digits[4:8]), int(digits[2:4]), int(digits[0:2]))
        except Exception:
            return None
    if len(digits) == 6:
        try:
            return date(2000 + int(digits[4:6]), int(digits[2:4]), int(digits[0:2]))
        except Exception:
            return None
    return None


def _dias_atraso(prevista: date | None, hoje: date | None = None) -> int:
    if prevista is None:
        return 0
    ref = hoje or date.today()
    return max(0, int((ref - prevista).days))


def _write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    ensure_dirs()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def parse_excel_081(path: Path | str) -> list[dict[str, Any]]:
    p = Path(path)
    if not p.is_file():
        return []
    suf = p.suffix.lower()
    if suf in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx_081(p)
    return _parse_text_081(p)


def _row_to_item(cells: list[Any], metric: dict[str, int], hoje: date) -> dict[str, Any] | None:
    ctrc = _clean(_cell_by_letter(cells, COL_CTRC))
    if not ctrc:
        return None
    cliente = _clean(_cell_by_letter(cells, COL_CLIENTE))
    raw_prev = _cell_by_letter(cells, COL_PREVISTA)
    prevista = _parse_date(raw_prev)
    cidade = _clean(_cell_by_letter(cells, COL_CIDADE))
    uf = _clean(_cell_by_letter(cells, COL_UF))
    ocorr = _clean(_cell_by_letter(cells, COL_OCORRENCIA))
    frete = 0.0
    peso = 0.0
    if "frete" in metric and metric["frete"] < len(cells):
        frete = _num(cells[metric["frete"]])
    if "peso" in metric and metric["peso"] < len(cells):
        peso = _num(cells[metric["peso"]])
    cidade_uf = f"{cidade}/{uf}".strip("/") if (cidade or uf) else ""
    if isinstance(raw_prev, (date, datetime)) and prevista:
        data_prevista = prevista.strftime("%d/%m/%Y")
    else:
        data_prevista = _clean(raw_prev)
    return {
        "ctrc": ctrc,
        "cliente": cliente,
        "data_prevista": data_prevista,
        "dias_atraso": _dias_atraso(prevista, hoje),
        "cidade": cidade,
        "uf": uf,
        "cidade_uf": cidade_uf,
        "ultima_ocorrencia": ocorr,
        "frete": frete,
        "peso": peso,
    }


def _parse_xlsx_081(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter) or ())
        except StopIteration:
            return []
        metric = _find_metric_cols(headers)
        hoje = date.today()
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if not row:
                continue
            item = _row_to_item(list(row), metric, hoje)
            if item:
                out.append(item)
        return out
    finally:
        wb.close()


def _parse_text_081(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise RuntimeError("081: encoding inválido")
    sample = text.splitlines()[:5]
    sep = ";" if sum(l.count(";") for l in sample) >= sum(l.count(",") for l in sample) else ","
    reader = csv.reader(text.splitlines(), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []
    metric = _find_metric_cols(rows[0])
    hoje = date.today()
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row:
            continue
        item = _row_to_item(list(row), metric, hoje)
        if item:
            out.append(item)
    return out


def analyze_reports_081(
    files: list[str] | list[Path] | str | Path | None,
    *,
    periodo: str = "",
    on_status=None,
) -> dict[str, Any]:
    status = on_status or (lambda _m: None)
    paths: list[Path] = []
    if isinstance(files, (str, Path)):
        paths = [Path(files)]
    elif files:
        paths = [Path(f) for f in files]

    rows: list[dict[str, Any]] = []
    for p in paths:
        if not p.is_file():
            continue
        status(f"[081] parse {p.name}")
        rows.extend(parse_excel_081(p))

    qtd = len(rows)
    frete = sum(float(r.get("frete") or 0) for r in rows)
    peso = sum(float(r.get("peso") or 0) for r in rows)

    top_cte = sorted(
        rows,
        key=lambda r: (int(r.get("dias_atraso") or 0), str(r.get("ctrc") or "")),
        reverse=True,
    )[:TOP_N]
    top_cte_out = [
        {
            "ctrc": r.get("ctrc") or "",
            "cliente": r.get("cliente") or "",
            "dias_atraso": int(r.get("dias_atraso") or 0),
            "cidade_uf": r.get("cidade_uf") or "",
            "ultima_ocorrencia": r.get("ultima_ocorrencia") or "",
        }
        for r in top_cte
    ]

    by_cli: dict[str, dict[str, Any]] = {}
    for r in rows:
        cli = str(r.get("cliente") or "").strip() or "(sem cliente)"
        slot = by_cli.setdefault(
            cli, {"cliente": cli, "dias_max": 0, "qtd": 0, "frete": 0.0, "peso": 0.0}
        )
        slot["qtd"] += 1
        slot["dias_max"] = max(int(slot["dias_max"]), int(r.get("dias_atraso") or 0))
        slot["frete"] += float(r.get("frete") or 0)
        slot["peso"] += float(r.get("peso") or 0)
    top_cli = sorted(
        by_cli.values(),
        key=lambda x: (int(x["dias_max"]), int(x["qtd"])),
        reverse=True,
    )[:TOP_N]

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumo = {
        "atualizado": agora,
        "periodo": periodo or "",
        "qtd": qtd,
        "frete": frete,
        "frete_fmt": _fmt_money(frete),
        "peso": peso,
        "peso_fmt": _fmt_int(peso),
    }

    ensure_dirs()
    _write_csv(ITENS_081_CSV, ITEM_FIELDS, rows)
    _write_csv(RESUMO_081_CSV, RESUMO_FIELDS, [resumo])
    _write_csv(TOP_CTE_081_CSV, TOP_CTE_FIELDS, top_cte_out)
    _write_csv(TOP_CLIENTE_081_CSV, TOP_CLIENTE_FIELDS, top_cli)

    payload = {
        "ok": True,
        "programa": "081",
        "view": "sem_saida",
        "resumo": resumo,
        "top_cte": top_cte_out,
        "top_cliente": top_cli,
        "itens": rows,
        "total": qtd,
    }
    LAST_081_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    status(f"[081] OK · CTRCs={qtd} frete={resumo['frete_fmt']} peso={resumo['peso_fmt']}")
    return payload
