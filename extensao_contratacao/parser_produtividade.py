"""Lê a planilha PRODUTIVIDADE CONTRATAÇÃO.xlsx → linhas tipadas."""
from __future__ import annotations

import os
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXCEL_FILENAME = "PRODUTIVIDADE CONTRATAÇÃO.xlsx"

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "data": ("DATA",),
    "motorista": ("MOTORISTA",),
    "cavalo": ("CAVALO", "PLACA", "PLACA CAVALO"),
    "carreta1": ("CARRETA1", "CARRETA 1", "CARRETA"),
    "carreta2": ("CARRETA2", "CARRETA 2", "CARRETA2"),
    "manifesto": (
        "MANIFESTO / ROMANEIO",
        "MANIFESTO/ROMANEIO",
        "NUMERO MANIFESTO / ROMANEIO",
        "NÚMERO MANIFESTO / ROMANEIO",
        "MANIFESTO",
    ),
    "proprietario": ("PROPRIETARIO", "PROPRIETÁRIO", "PROPRIETARIO"),
    "origem": ("ORIGEM",),
    "destino": ("DESTINO",),
    "frete_fechado": ("FRETE FECHADO", "FRETE"),
    "status": ("STATUS", "SITUAÇÃO", "SITUACAO"),
}


def norm_placa(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def desktop_dir() -> Path:
    """Área de Trabalho do usuário atual (Desktop ou OneDrive\\Desktop)."""
    home = Path.home()
    userprofile = Path(os.environ.get("USERPROFILE") or str(home))
    candidates = [
        home / "Desktop",
        userprofile / "Desktop",
        home / "OneDrive" / "Desktop",
        userprofile / "OneDrive" / "Desktop",
        home / "OneDrive - Pessoal" / "Desktop",
        userprofile / "OneDrive - Pessoal" / "Desktop",
    ]
    for path in candidates:
        if path.is_dir():
            return path
    return userprofile / "Desktop"


def resolve_produtividade_xlsx(configured: str | Path | None = None) -> Path:
    """
    Sempre resolve a planilha na Área de Trabalho.

    - Ignora pastas configuradas fora do Desktop.
    - Aceita só o nome do arquivo (ou vazio → PRODUTIVIDADE CONTRATAÇÃO.xlsx).
    """
    name = EXCEL_FILENAME
    raw = str(configured or "").strip()
    if raw:
        candidate_name = Path(raw).name.strip()
        if candidate_name.lower().endswith((".xlsx", ".xlsm", ".xls")):
            name = candidate_name
    return desktop_dir() / name


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _parse_money(raw: Any) -> float:
    if isinstance(raw, (int, float)):
        return float(raw)
    text = _clean(raw).replace(" ", "").replace("R$", "").replace("$", "")
    if not text or text in {"-", "."}:
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        if text.count(".") == 1 and len(text.split(".")[-1]) == 2:
            pass
        else:
            text = text.replace(".", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _header_key(cell: Any) -> str:
    return re.sub(r"\s+", " ", _clean(cell).upper())


def _find_header_row(rows: list[tuple[Any, ...]], max_scan: int = 40) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows[:max_scan]):
        labels = {_header_key(c): i for i, c in enumerate(row) if _clean(c)}
        if "DATA" not in labels:
            continue
        if not any(k in labels for k in ("CAVALO", "PLACA", "PLACA CAVALO")):
            continue
        mapping: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in labels:
                    mapping[field] = labels[alias]
                    break
        if "data" in mapping and "cavalo" in mapping:
            return idx, mapping
    raise RuntimeError("Cabeçalho DATA/CAVALO não encontrado na planilha.")


def _month_sheet_name(hoje: date | None = None) -> str:
    d = hoje or date.today()
    return f"{d.month:02d} {d.year}"


def pick_sheet_name(workbook, *, prefer: str = "") -> str:
    names = list(workbook.sheetnames)
    if prefer and prefer in names:
        return prefer
    month = _month_sheet_name()
    if month in names:
        return month
    if "PLANILHA MAE" in names:
        return "PLANILHA MAE"
    return names[0]


def is_cancelado(status: str, manifesto: str = "") -> bool:
    blob = f"{status} {manifesto}".upper()
    return "CANCEL" in blob


def read_produtividade_xlsx(
    path: Path | str,
    *,
    sheet: str = "",
    only_month_to_today: bool = True,
    only_ontem_hoje: bool = False,
    hoje: date | None = None,
) -> dict[str, Any]:
    """
    Lê a planilha e devolve linhas tipadas.

    Por padrão: mês referente (dia 1 até hoje) na aba do mês vigente.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {path}")

    day = hoje or date.today()
    ontem = day - timedelta(days=1)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = pick_sheet_name(wb, prefer=sheet.strip())
        ws = wb[sheet_name]
        raw_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    header_idx, cols = _find_header_row(raw_rows)
    month_start = date(day.year, day.month, 1)
    rows: list[dict[str, Any]] = []
    skipped_cancel = 0
    skipped_date = 0

    for raw in raw_rows[header_idx + 1 :]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue

        def cell(key: str) -> Any:
            i = cols.get(key)
            if i is None or i >= len(raw):
                return ""
            return raw[i]

        placa = norm_placa(cell("cavalo"))
        if not placa:
            continue

        status = _clean(cell("status"))
        manifesto = _clean(cell("manifesto"))
        if is_cancelado(status, manifesto):
            skipped_cancel += 1
            continue

        dt = _as_date(cell("data"))
        if only_ontem_hoje:
            if dt is None or dt < ontem or dt > day:
                skipped_date += 1
                continue
        elif only_month_to_today:
            if dt is None:
                skipped_date += 1
                continue
            if dt < month_start or dt > day:
                skipped_date += 1
                continue

        carreta = norm_placa(cell("carreta1")) or _clean(cell("carreta1")).upper()
        if carreta in {"-", "—", "N/A", "NA"}:
            carreta = ""
        carreta2 = norm_placa(cell("carreta2"))
        if carreta2 and carreta2 not in {"-", "—"}:
            carreta = f"{carreta}+{carreta2}" if carreta else carreta2

        rows.append(
            {
                "data": dt.strftime("%d/%m/%Y") if dt else "",
                "data_iso": dt.isoformat() if dt else "",
                "motorista": _clean(cell("motorista")),
                "placa": placa,
                "carreta": carreta,
                "manifesto": manifesto,
                "propriedade": _clean(cell("proprietario")),
                "origem": _clean(cell("origem")),
                "destino": _clean(cell("destino")),
                "custo": round(_parse_money(cell("frete_fechado")), 2),
                "status": status or "CONTRATADO",
                "fonte": path.name,
                "aba": sheet_name,
            }
        )

    return {
        "ok": True,
        "path": str(path),
        "sheet": sheet_name,
        "rows": rows,
        "total": len(rows),
        "skipped_cancel": skipped_cancel,
        "skipped_date": skipped_date,
        "periodo_ref": (
            f"{month_start.strftime('%d/%m/%Y')} a {day.strftime('%d/%m/%Y')}"
            if only_month_to_today or not only_ontem_hoje
            else f"{ontem.strftime('%d/%m/%Y')} a {day.strftime('%d/%m/%Y')}"
        ),
    }
