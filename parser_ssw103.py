from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime, time
from pathlib import Path
from typing import Any

from config import BASE_DIR, ensure_dirs

CACHE_DIR = BASE_DIR / "data" / "cache"
COLETAS_103_CSV = CACHE_DIR / "coletas_103.csv"
RESUMO_103_CSV = CACHE_DIR / "resumo_103.csv"
LAST_103_JSON = CACHE_DIR / "last_run_103.json"

# Colunas Excel (1-based) no CSV SSW (com coluna tipo "1"/"3" na A):
# AE=COLETAR_DATA AF=COLETAR_HORA AI=SITUACAO AK=PLACA AL=CARRETA AN=MOTORISTA
COL_AE = 31  # COLETAR_DATA (ref)
COL_AF = 32  # COLETAR_HORA
COL_AI = 35  # SITUACAO (atual: CADASTRADA/COMANDADA/COLETADA)
COL_AK = 37  # PLACA_CAVALO
COL_AL = 38  # PLACA_CARRETA
COL_AN = 40  # MOTORISTA

# Headers preferidos (CSV ssw0166)
HDR_SITUACAO = "SITUACAO"
HDR_HORA = "COLETAR_HORA"
HDR_DATA_COLETAR = "COLETAR_DATA"
HDR_PLACA = "PLACA_CAVALO"
HDR_CARRETA = "PLACA_CARRETA"
HDR_MOTORISTA = "MOTORISTA"
HDR_UNIDADE = "UNIDADE"
HDR_NUMERO = "NUMERO_COLETA"


@dataclass
class Coleta103:
    coleta_id: str
    situacao_atual: str = ""  # texto bruto AE
    status_ace: str = ""  # parado | em_rota | realizada | cancelada | outro
    hora: str = ""
    hora_antes_meio_dia: bool = False
    cadastrada_ref: str = ""  # AI
    placa: str = ""
    placa_carreta: str = ""
    motorista: str = ""
    extras: dict[str, str] = field(default_factory=dict)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%H:%M")
        except Exception:
            pass
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _col_letter(idx: int) -> str:
    letters = ""
    n = idx
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _parse_hora(value: Any) -> tuple[str, time | None]:
    """Retorna (texto HH:MM, time|None)."""
    if value is None or value == "":
        return "", None
    if isinstance(value, datetime):
        return value.strftime("%H:%M"), value.time()
    if isinstance(value, time):
        return value.strftime("%H:%M"), value
    # Excel serial fraction of day
    if isinstance(value, (int, float)) and 0 <= float(value) < 1.5:
        total = int(round(float(value) * 24 * 60)) % (24 * 60)
        hh, mm = divmod(total, 60)
        return f"{hh:02d}:{mm:02d}", time(hh, mm)
    text = _clean(value)
    m = re.search(r"(\d{1,2}):(\d{2})", text)
    if m:
        hh, mm = int(m.group(1)), int(m.group(2))
        if 0 <= hh < 24 and 0 <= mm < 60:
            return f"{hh:02d}:{mm:02d}", time(hh, mm)
    return text, None


def mapear_status_103(situacao_ae: str, cadastrada_ai: str = "", hora_t: time | None = None) -> str:
    """
    Padrao ACE 103:
      cadastrada → parado
      comandada  → em_rota
      coletada   → realizada
    AE = situacao atual (fonte principal).
    Se AE vazio e AI indica cadastrada e AF < 12:00 → parado.
    """
    ae = (situacao_ae or "").upper()
    ai = (cadastrada_ai or "").upper()

    def classify(text: str) -> str:
        if not text:
            return ""
        if "CANCEL" in text:
            return "cancelada"
        if "COLET" in text or "REALIZ" in text:
            return "realizada"
        if "COMAND" in text or "ROTA" in text or "TRANSITO" in text or "TRÂNSITO" in text:
            return "em_rota"
        if "CADASTR" in text or "PARADO" in text or "PEND" in text:
            return "parado"
        return ""

    status = classify(ae)
    if status:
        return status

    # Ref AI + antes do meio-dia
    if hora_t is not None and hora_t < time(12, 0):
        status = classify(ai)
        if status:
            return status
        if ai.strip():
            return "parado"

    status = classify(ai)
    return status or "outro"


def _cell(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _guess_coleta_id(ws, row: int) -> str:
    """Tenta unidade+numero nas primeiras colunas."""
    parts: list[str] = []
    for col in range(1, 12):
        val = _clean(_cell(ws, row, col))
        if not val:
            continue
        if re.fullmatch(r"[A-Za-z]{3}", val):
            parts.append(val.upper())
        elif re.fullmatch(r"\d{5,8}", val):
            parts.append(val)
        elif re.fullmatch(r"[A-Za-z]{3}\s*\d{5,8}", val):
            return re.sub(r"\s+", "", val.upper())
        if len(parts) >= 2:
            return f"{parts[0]}{parts[1]}"
    # fallback: procura padrao SPO 123456 em qualquer celula da linha (A-P)
    for col in range(1, 16):
        val = _clean(_cell(ws, row, col))
        m = re.search(r"([A-Za-z]{3})\s*(\d{5,8})", val)
        if m:
            return f"{m.group(1).upper()}{m.group(2)}"
    return f"LINHA{row}"


def _norm_header(name: str) -> str:
    return re.sub(r"\s+", "_", _clean(name).upper())


def _read_csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    raw = path.read_bytes()
    text = raw.decode("cp1252", errors="replace")
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")
    # detecta separador
    sample = text[:4000]
    sep = ";" if sample.count(";") >= sample.count(",") else ","
    rows = [line.split(sep) for line in text.splitlines() if line.strip()]
    if not rows:
        return [], []
    # Header principal: linha que contem SITUACAO + UNIDADE
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        joined = ";".join(_norm_header(c) for c in row)
        if "SITUACAO" in joined and "UNIDADE" in joined and "NUMERO" in joined:
            header_idx = i
            break
    header = [_norm_header(c) for c in rows[header_idx]]
    data = rows[header_idx + 1 :]
    return header, data


def _idx(header: list[str], *names: str) -> int | None:
    """Retorna o indice do primeiro nome encontrado (ordem de prioridade em names)."""
    normalized = [_norm_header(h) for h in header]
    for name in names:
        want = _norm_header(name)
        if want in normalized:
            return normalized.index(want)
    return None


def parse_ssw103_csv(path: Path | str) -> list[Coleta103]:
    """CSV/sswweb do ssw0166 (Mostrar em = E)."""
    file_path = Path(path)
    header, data = _read_csv_rows(file_path)
    if not header:
        return []

    i_sit = _idx(header, HDR_SITUACAO)
    i_hora = _idx(header, HDR_HORA, "HORA_INCLUSAO")
    i_ai_ref = i_sit  # AI = SITUACAO (cadastrada/comandada/coletada)
    i_placa = _idx(header, HDR_PLACA, "PLACA")
    i_carr = _idx(header, HDR_CARRETA)
    i_mot = _idx(header, HDR_MOTORISTA)
    i_uni = _idx(header, HDR_UNIDADE)
    i_num = _idx(header, HDR_NUMERO, "NUMERO")
    i_data = _idx(header, HDR_DATA_COLETAR)

    # Fallback posicional (letras Excel) se header incompleto
    def col(row: list[str], idx: int | None, excel_1based: int) -> str:
        if idx is not None and 0 <= idx < len(row):
            return _clean(row[idx])
        pos = excel_1based - 1
        if 0 <= pos < len(row):
            return _clean(row[pos])
        return ""

    coletas: list[Coleta103] = []
    seen: set[str] = set()
    for n, row in enumerate(data, start=1):
        # pula sub-headers (tipo 2 / linhas curtas / UNIDADE no campo unidade)
        if len(row) < 10:
            continue
        uni = col(row, i_uni, 2)
        if uni.upper() in {"", "UNIDADE"}:
            continue
        num = col(row, i_num, 3)
        if not re.search(r"\d", num):
            continue

        ae_data = col(row, i_data, COL_AE)  # COLETAR_DATA
        ai = col(row, i_ai_ref, COL_AI)  # SITUACAO
        hora_txt, hora_t = _parse_hora(col(row, i_hora, COL_AF))
        placa = col(row, i_placa, COL_AK)
        carreta = col(row, i_carr, COL_AL)
        motorista = col(row, i_mot, COL_AN)

        # AE (data coletar) nao e status; situacao atual = AI/SITUACAO
        status = mapear_status_103(ai, ai, hora_t)
        num_digits = re.sub(r"\D", "", num)
        cid = f"{uni.upper()}{num_digits}"
        rec = Coleta103(
            coleta_id=cid,
            situacao_atual=ai,
            status_ace=status,
            hora=hora_txt,
            hora_antes_meio_dia=bool(hora_t and hora_t < time(12, 0)),
            cadastrada_ref=ai,
            placa=placa,
            placa_carreta=carreta,
            motorista=motorista,
            extras={
                "linha": str(n),
                "AE_COLETAR_DATA": ae_data,
                "AF": hora_txt,
                "AI": ai,
            },
        )
        if cid in seen:
            continue
        seen.add(cid)
        coletas.append(rec)
    return coletas


def parse_ssw103_excel(path: Path | str) -> list[Coleta103]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "openpyxl nao instalado. Rode: pip install openpyxl"
        ) from error

    file_path = Path(path)
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    coletas: list[Coleta103] = []
    seen: set[str] = set()

    for row in range(2, (ws.max_row or 1) + 1):
        # AE=data coletar; AI=situacao atual; AF=hora
        ae = _clean(_cell(ws, row, COL_AE))
        ai = _clean(_cell(ws, row, COL_AI))
        hora_txt, hora_t = _parse_hora(_cell(ws, row, COL_AF))
        placa = _clean(_cell(ws, row, COL_AK))
        carreta = _clean(_cell(ws, row, COL_AL))
        motorista = _clean(_cell(ws, row, COL_AN))

        if not any([ai, ae, hora_txt, placa, carreta, motorista]):
            if not any(_clean(_cell(ws, row, c)) for c in range(1, 8)):
                continue

        cid = _guess_coleta_id(ws, row)
        status = mapear_status_103(ai, ai, hora_t)
        rec = Coleta103(
            coleta_id=cid,
            situacao_atual=ai or ae,
            status_ace=status,
            hora=hora_txt,
            hora_antes_meio_dia=bool(hora_t and hora_t < time(12, 0)),
            cadastrada_ref=ai,
            placa=placa,
            placa_carreta=carreta,
            motorista=motorista,
            extras={
                "linha_excel": str(row),
                "AE": ae,
                "AF": hora_txt,
                "AI": ai,
            },
        )
        if cid in seen:
            idx = next(i for i, c in enumerate(coletas) if c.coleta_id == cid)
            score_new = sum(bool(x) for x in (ai, placa, motorista))
            score_old = sum(
                bool(x)
                for x in (coletas[idx].situacao_atual, coletas[idx].placa, coletas[idx].motorista)
            )
            if score_new >= score_old:
                coletas[idx] = rec
            continue
        seen.add(cid)
        coletas.append(rec)

    wb.close()
    return coletas


def parse_ssw103(path: Path | str) -> list[Coleta103]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    name = file_path.name.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return parse_ssw103_excel(file_path)
    # CSV / .sswweb (SSW "Excel" da 103)
    if suffix in {".csv", ".sswweb", ".txt"} or "ssw0166" in name or name.startswith("csv"):
        return parse_ssw103_csv(file_path)
    # tenta CSV primeiro (mais comum na 103), depois excel
    try:
        rows = parse_ssw103_csv(file_path)
        if rows:
            return rows
    except Exception:
        pass
    return parse_ssw103_excel(file_path)


def contar_status_103(coletas: list[Coleta103]) -> dict[str, int]:
    out = {
        "total": 0,
        "parado": 0,
        "em_rota": 0,
        "realizada": 0,
        "cancelada": 0,
        "outro": 0,
    }
    for c in coletas:
        out["total"] += 1
        key = c.status_ace if c.status_ace in out else "outro"
        out[key] += 1
    return out


def coleta103_to_row(c: Coleta103) -> dict[str, Any]:
    return {
        "coleta_id": c.coleta_id,
        "situacao_atual": c.situacao_atual,
        "status_ace": c.status_ace,
        "hora": c.hora,
        "hora_antes_meio_dia": "S" if c.hora_antes_meio_dia else "N",
        "cadastrada_ref_AI": c.cadastrada_ref,
        "placa": c.placa,
        "placa_carreta": c.placa_carreta,
        "motorista": c.motorista,
    }


COLETA_103_FIELDS = list(coleta103_to_row(Coleta103(coleta_id="X")).keys())
RESUMO_103_FIELDS = ["periodo", "total", "parado", "em_rota", "realizada", "cancelada", "outro"]


def save_cache_103(
    coletas: list[Coleta103],
    *,
    source_file: str = "",
    periodo: str = "",
) -> dict[str, Any]:
    ensure_dirs()
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    rows = [coleta103_to_row(c) for c in coletas]
    totais = contar_status_103(coletas)
    resumo = [{"periodo": periodo or "PERIODO", **totais}]

    with COLETAS_103_CSV.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLETA_103_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    with RESUMO_103_CSV.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=RESUMO_103_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(resumo)

    meta = {
        "ok": True,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "source_file": source_file,
        "periodo": periodo,
        "lote": len(coletas),
        "totais": totais,
        "modelo": (
            "103 CSV/ssw0166: AI/SITUACAO=atual, AF/COLETAR_HORA=hora, "
            "AE/COLETAR_DATA=ref, AK/AL/AN=placa/carreta/motorista. "
            "parado←cadastrada, em_rota←comandada, realizada←coletada."
        ),
        "colunas": {
            "AE": _col_letter(COL_AE),
            "AF": _col_letter(COL_AF),
            "AI": _col_letter(COL_AI),
            "AK": _col_letter(COL_AK),
            "AL": _col_letter(COL_AL),
            "AN": _col_letter(COL_AN),
        },
        "paths": {"coletas": str(COLETAS_103_CSV), "resumo": str(RESUMO_103_CSV)},
    }
    LAST_103_JSON.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
    return meta


def analyze_report_103(path: Path | str, *, periodo: str = "") -> dict[str, Any]:
    file_path = Path(path)
    coletas = parse_ssw103(file_path)
    meta = save_cache_103(coletas, source_file=str(file_path), periodo=periodo)
    meta["records"] = [asdict(c) for c in coletas]
    return meta
