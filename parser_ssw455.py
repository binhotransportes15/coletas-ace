"""Parser Excel SSW 455 — Fretes Expedidos/Recebidos (Emissão).

Agrega KPIs do painel com colunas fixas do Excel:
  G=data emissão · H=hora emissão · I=data autorização · J=hora autorização
  K=login · BD=peso · BE=cubagem · BF=volumes
  BI=tipo baixa · BN=valor mercadoria · BQ=frete
  DIA / NOITE (hora de emissão, col. H) · Expedidores (coluna K)
  Pendentes = com emissão sem autorização · Finalizados = com autorização
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from config import BASE_DIR, ensure_dirs
from dates import MESES_PT

CACHE_DIR = BASE_DIR / "data" / "cache"
EMISSOES_455_CSV = CACHE_DIR / "emissoes_455.csv"
RESUMO_455_CSV = CACHE_DIR / "resumo_455.csv"
EXPEDIDORES_455_CSV = CACHE_DIR / "expedidores_455.csv"
HORAS_455_CSV = CACHE_DIR / "horas_455.csv"
LAST_455_JSON = CACHE_DIR / "last_run_455.json"

# DIA = 06:00–17:59 · NOITE = 18:00–05:59 (hora de emissão)
DIA_START = 6
DIA_END = 18  # exclusive

# Excel 455: coluna K = Login (1-based) — base dos expedidores
COL_K_LOGIN = 11

# Colunas fixas do Excel 455 (letras → métricas do painel)
# Ordem real SSW: G/H emissão · I/J autorização · BD=peso · BE=cubagem · BF=volumes · BN=merc · BQ=frete
COL_FIXED_455: dict[str, str] = {
    "data_emissao": "G",
    "hora": "H",  # hora de emissão (picos / dia-noite)
    "data_autorizacao": "I",
    "hora_autorizacao": "J",
    "login": "K",
    "peso": "BD",
    "cubagem": "BE",
    "volumes": "BF",
    "cancelado": "BI",
    "valor_mercadoria": "BN",
    "frete": "BQ",
}

# Ranking do painel: só estes logins · exibição = nome completo (direita)
LOGIN_NOME_455: dict[str, str] = {
    "l.marque": "Lidiane Marques",
    "eloi": "Emerson Eloi",
    "g.fagund": "Gabriel Fagundes",
    "f.silva": "Flaviana Eneas",
    "e.beliza": "Edson Belizario",
    "m.cordei": "Maria Eduarda",
    "anderson": "Anderson Vieira",
    "s.silva": "Simone Silva",
    "j.rodrig": "Julia Rodrigues",
    "portal": "Portal",
    "g.nascim": "Gabriely Nascimento",
    "m.neres": "Mayara Neres",
}


def _col_letter_to_idx(letters: str) -> int:
    """Excel 'A'→0, 'H'→7, 'BD'→55, 'BQ'→68."""
    n = 0
    for ch in str(letters or "").strip().upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - ord("A") + 1)
    return max(0, n - 1)


def _cell_by_letter(row: tuple[Any, ...] | list[Any], letters: str) -> Any:
    idx = _col_letter_to_idx(letters)
    if idx >= len(row):
        return ""
    return row[idx]

EMISSAO_FIELDS = [
    "ctrc",
    "data_emissao",
    "hora_emissao",
    "data_autorizacao",
    "hora_autorizacao",
    "expedidor",
    "frete",
    "valor_mercadoria",
    "peso",
    "volumes",
    "cubagem",
    "cancelado",
    "unidade",
    "liquidacao",
]

RESUMO_FIELDS = [
    "periodo",
    "mes",
    "atualizado",
    "ctes",
    "peso",
    "peso_fmt",
    "valor_mercadoria",
    "valor_mercadoria_fmt",
    "volumes",
    "cubagem",
    "cubagem_fmt",
    "frete",
    "frete_fmt",
    "dia",
    "noite",
    "cancelados",
    "pendentes",
    "finalizados",
]

EXPEDIDOR_FIELDS = ["nome", "nome_exibicao", "qtd", "pct"]
HORA_FIELDS = ["hora", "label", "qtd"]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            pass
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    try:
        n = float(text or 0)
        return -n if neg else n
    except Exception:
        return 0.0


def _fmt_int(n: float | int) -> str:
    return f"{int(round(float(n) or 0)):,}".replace(",", ".")


def _fmt_dec(n: float, places: int = 2) -> str:
    s = f"{float(n or 0):,.{places}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_money(n: float) -> str:
    return f"R$ {_fmt_dec(n, 2)}"


def _write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    ensure_dirs()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def _norm_header(h: Any) -> str:
    t = _clean(h).lower()
    t = t.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e")
    t = t.replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ê", "e")
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def _map_headers(headers: list[Any], *, apply_fixed: bool = True) -> dict[str, int]:
    """Mapeia nomes canônicos → índice 0-based (+ letras fixas só no Excel largo)."""
    norms = [_norm_header(h) for h in headers]
    out: dict[str, int] = {}

    def pick(keys: tuple[str, ...], *aliases: str) -> None:
        for i, n in enumerate(norms):
            if not n:
                continue
            for a in aliases:
                if a in n:
                    for k in keys:
                        out.setdefault(k, i)
                    return

    pick(("ctrc",), "serie numero", "serie/numero", "nr cte", "n cte", "ctrc", "conhecimento", "cte")
    # CSV slim: coluna "EMISS" (não confundir com "unidade emissora")
    for i, n in enumerate(norms):
        if n == "emiss" or n.startswith("emiss ") or "data de emissao" in n or n == "data emissao" or n == "dt emissao":
            out.setdefault("data_emissao", i)
            break
    pick(("hora",), "hora de emissao", "hora emissao", "hr emissao")
    pick(("data_autorizacao",), "data de autorizacao", "data autorizacao", "dt autoriz")
    pick(("hora_autorizacao",), "hora de autorizacao", "hora autorizacao", "hr autoriz", "hora autoriz")
    pick(("unidade",), "unidade emissora", "unidade cobranca", "unidade responsavel", "unidade", "filial", "sigla")
    pick(("liquidacao",), "liquidacao", "liq ")
    pick(("peso",), "peso real", "peso total", "peso kg", "peso")
    # cubagem real; "peso cubado" no slim NÃO é m³ — só usa se não houver cubagem
    pick(("cubagem",), "cubagem", "m3", "metro cub")
    if "cubagem" not in out:
        for i, n in enumerate(norms):
            if "peso cubado" in n:
                # slim: sem m³ — deixa 0 (não mapear peso cubado como cubagem)
                break
    pick(("volumes",), "quantidade de volume", "qtde volume", "qtd volume", "volumes", "volume")
    pick(("valor_mercadoria",), "valor da mercadoria", "valor mercadoria", "vl mercadoria", "valmercad")
    pick(("frete",), "valor do frete", "valor frete", "frete total", "vl frete", "valfrete", "frete")
    pick(("cancelado",), "cancelado", "anulado", "tipo de baixa")
    pick(("expedidor",), "login", "usuario", "usuário", "expedidor", "conferente")

    # Letras fixas só quando a linha tem largura de Excel completo (≥ coluna BQ)
    min_wide = _col_letter_to_idx(COL_FIXED_455["frete"]) + 1
    if apply_fixed and len(headers) >= min_wide:
        out["data_emissao"] = _col_letter_to_idx(COL_FIXED_455["data_emissao"])
        out["hora"] = _col_letter_to_idx(COL_FIXED_455["hora"])
        out["data_autorizacao"] = _col_letter_to_idx(COL_FIXED_455["data_autorizacao"])
        out["hora_autorizacao"] = _col_letter_to_idx(COL_FIXED_455["hora_autorizacao"])
        out["expedidor"] = _col_letter_to_idx(COL_FIXED_455["login"])
        out["peso"] = _col_letter_to_idx(COL_FIXED_455["peso"])
        out["cubagem"] = _col_letter_to_idx(COL_FIXED_455["cubagem"])
        out["volumes"] = _col_letter_to_idx(COL_FIXED_455["volumes"])
        out["cancelado"] = _col_letter_to_idx(COL_FIXED_455["cancelado"])
        out["valor_mercadoria"] = _col_letter_to_idx(COL_FIXED_455["valor_mercadoria"])
        out["frete"] = _col_letter_to_idx(COL_FIXED_455["frete"])
    return out


def _is_slim_headers(headers: list[Any]) -> bool:
    """CSV reduzido do 455 (CTRC;EMISS;…;VALFRETE) — sem colunas G/H/K/BQ."""
    norms = " | ".join(_norm_header(h) for h in headers)
    if "valfrete" in norms or re.search(r"\bemiss\b", norms):
        return True
    return len(headers) < _col_letter_to_idx(COL_FIXED_455["frete"]) + 1


def _is_junk_row(rec: dict[str, Any], *, require_login: bool = True) -> bool:
    """Descarta cabeçalho residual / linhas sem movimento."""
    login = str(rec.get("expedidor") or "").strip().lower()
    ctrc = str(rec.get("ctrc") or "").strip().lower()
    if require_login and login in {"", "login", "usuario", "usuário", "expedidor", "conferente"}:
        return True
    if any(x in ctrc for x in ("serie", "ct-e", "ctrc", "conhecimento", "numero")):
        return True
    if ctrc in {"0", "0.0", "-"}:
        if float(rec.get("peso") or 0) <= 0 and float(rec.get("frete") or 0) <= 0:
            return True
    # CSV slim: login às vezes vem como valor monetário por mapeamento errado antigo
    if login and re.fullmatch(r"[\d.,]+", login.replace(" ", "")):
        if float(rec.get("frete") or 0) <= 0 and float(rec.get("peso") or 0) <= 0:
            return True
    return False


def _parse_hora_value(val: Any) -> int | None:
    """Extrai hora 0–23 de célula Excel/texto."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.hour
    if hasattr(val, "hour") and not isinstance(val, str):
        try:
            return int(val.hour)
        except Exception:
            pass
    # Excel time as fraction of day
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            f = float(val)
            if 0 <= f < 1.5:  # serial time
                return max(0, min(23, int(f * 24) % 24))
            if 0 <= f <= 23:
                return int(f)
        except Exception:
            pass
    text = _clean(val)
    if not text:
        return None
    m = re.search(r"(\d{1,2}):(\d{2})", text)
    if m:
        return max(0, min(23, int(m.group(1))))
    m = re.search(r"\b(\d{1,2})h", text, re.I)
    if m:
        return max(0, min(23, int(m.group(1))))
    if re.fullmatch(r"\d{1,2}", text):
        return max(0, min(23, int(text)))
    return None


def _cell_raw(row: tuple[Any, ...] | list[Any], colmap: dict[str, int], key: str) -> Any:
    idx = colmap.get(key)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def _fmt_hora_hhmm(hora: int | None, raw: Any = None) -> str:
    if hora is not None:
        # preserva minutos se o texto original tiver HH:MM
        text = _clean(raw)
        m = re.search(r"(\d{1,2}):(\d{2})", text)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return f"{hora:02d}:00"
    return ""


def _has_datetime_parts(*, data: str, hora_txt: str, hora_int: int | None) -> bool:
    return bool(str(data or "").strip()) or bool(str(hora_txt or "").strip()) or hora_int is not None


def _parse_hora(row: dict[str, Any], raw_row: tuple[Any, ...] | list[Any], colmap: dict[str, int]) -> int | None:
    """Compat: hora de emissão (coluna H)."""
    return _parse_hora_value(_cell_raw(raw_row, colmap, "hora"))


def _enrich_row_times(
    rec: dict[str, Any],
    raw_row: tuple[Any, ...] | list[Any],
    colmap: dict[str, int],
) -> None:
    """Preenche emissão / autorização e flags internas."""
    raw_he = _cell_raw(raw_row, colmap, "hora")
    raw_ha = _cell_raw(raw_row, colmap, "hora_autorizacao")
    raw_de = _cell_raw(raw_row, colmap, "data_emissao")
    raw_da = _cell_raw(raw_row, colmap, "data_autorizacao")

    h_em = _parse_hora_value(raw_he)
    h_au = _parse_hora_value(raw_ha)

    data_em = _clean(raw_de) or _clean(rec.get("data_emissao"))
    data_au = _clean(raw_da)
    hora_em = _fmt_hora_hhmm(h_em, raw_he)
    hora_au = _fmt_hora_hhmm(h_au, raw_ha)

    rec["data_emissao"] = data_em
    rec["hora_emissao"] = hora_em
    rec["data_autorizacao"] = data_au
    rec["hora_autorizacao"] = hora_au
    # picos / dia-noite: hora de emissão (filtro do relatório)
    rec["_hora"] = h_em if h_em is not None else h_au
    rec["_tem_emissao"] = _has_datetime_parts(data=data_em, hora_txt=hora_em, hora_int=h_em)
    rec["_tem_autorizacao"] = _has_datetime_parts(data=data_au, hora_txt=hora_au, hora_int=h_au)


def _is_cancelado(row: dict[str, str]) -> bool:
    """Coluna BI = cancelados (S/C/flag/texto)."""
    raw = (row.get("cancelado") or "").strip()
    if not raw:
        return False
    up = raw.upper()
    if up in {"S", "C", "1", "X", "SIM", "CANCELADO", "CANCELADA"}:
        return True
    if up in {"N", "0", "NAO", "NÃO", ""}:
        return False
    blob = " ".join(
        [
            raw,
            row.get("liquidacao") or "",
            row.get("ctrc") or "",
        ]
    ).lower()
    return bool(re.search(r"\bcancel|\banulad|\bsubstitu", blob))


def _norm_login(nome: str) -> str:
    return (nome or "").strip().rstrip("*＊").strip().lower()


def _is_expedidor(nome: str) -> bool:
    """Só logins da lista oficial do painel Emissão."""
    return _norm_login(nome) in LOGIN_NOME_455


def _display_expedidor(nome: str) -> str:
    """Nome completo no ranking (direita) — nunca o login."""
    key = _norm_login(nome)
    if key in LOGIN_NOME_455:
        return LOGIN_NOME_455[key]
    t = (nome or "").strip().rstrip("*＊").strip()
    return t.title() if t else "—"


def _login_from_row(
    row: tuple[Any, ...] | list[Any],
    colmap: dict[str, int],
    *,
    slim: bool = False,
) -> str:
    """Excel largo: coluna K. CSV slim: header login/conferente."""
    if not slim:
        idx_k = _col_letter_to_idx(COL_FIXED_455["login"])
        if idx_k < len(row):
            login_k = _clean(row[idx_k])
            if login_k:
                return login_k
    idx_h = colmap.get("expedidor")
    if idx_h is not None and idx_h < len(row):
        return _clean(row[idx_h])
    return ""


def _metrics_from_row(
    row: tuple[Any, ...] | list[Any],
    colmap: dict[str, int] | None = None,
    *,
    slim: bool = False,
) -> dict[str, Any]:
    """Lê métricas pelas letras fixas (Excel) ou pelo mapa de headers (CSV slim)."""
    if slim and colmap is not None:

        def g(key: str) -> Any:
            idx = colmap.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        return {
            "frete": _num(g("frete")),
            "valor_mercadoria": _num(g("valor_mercadoria")),
            "peso": _num(g("peso")),
            "volumes": _num(g("volumes")),
            "cubagem": _num(g("cubagem")),
            "cancelado": _clean(g("cancelado")),
        }
    return {
        "frete": _num(_cell_by_letter(row, COL_FIXED_455["frete"])),
        "valor_mercadoria": _num(_cell_by_letter(row, COL_FIXED_455["valor_mercadoria"])),
        "peso": _num(_cell_by_letter(row, COL_FIXED_455["peso"])),
        "volumes": _num(_cell_by_letter(row, COL_FIXED_455["volumes"])),
        "cubagem": _num(_cell_by_letter(row, COL_FIXED_455["cubagem"])),
        "cancelado": _clean(_cell_by_letter(row, COL_FIXED_455["cancelado"])),
    }


def parse_excel_455(path: Path | str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    p = Path(path)
    if not p.is_file():
        raise RuntimeError(f"455: arquivo inexistente: {p}")

    # .sswweb / csv / texto; ou .xlsx falso (SSW grava sswweb com extensão xlsx)
    suf = p.suffix.lower()
    head = b""
    try:
        head = p.read_bytes()[:8]
    except OSError:
        head = b""
    is_zip = head[:2] == b"PK"
    if suf in {".csv", ".sswweb", ".txt"} or not is_zip:
        return _parse_text_455(p)

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = None
        colmap: dict[str, int] = {}
        out: list[dict[str, Any]] = []
        for i, row in enumerate(rows_iter, start=1):
            if not row:
                continue
            if header_row is None:
                # detect header
                joined = " ".join(_clean(c).lower() for c in row if c is not None)
                if any(
                    k in joined
                    for k in ("ctrc", "cte", "frete", "emiss", "conhecimento", "peso")
                ):
                    header_row = row
                    colmap = _map_headers(list(row))
                    continue
                if i <= 5:
                    continue
                # sem header claro — aborta
                if i == 6:
                    raise RuntimeError("455: cabeçalho do Excel não reconhecido")
                continue

            def cell(key: str) -> Any:
                idx = colmap.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return row[idx]

            ctrc = _clean(cell("ctrc"))
            if not ctrc:
                continue
            up = ctrc.upper()
            if "CTRC" in up and len(ctrc) < 8:
                continue

            login = _login_from_row(row, colmap, slim=False)
            metrics = _metrics_from_row(row, colmap, slim=False)
            rec = {
                "ctrc": ctrc,
                "data_emissao": _clean(cell("data_emissao")),
                "hora_emissao": "",
                "data_autorizacao": "",
                "hora_autorizacao": "",
                "expedidor": login,  # coluna K · login
                "frete": metrics["frete"],
                "valor_mercadoria": metrics["valor_mercadoria"],
                "peso": metrics["peso"],
                "volumes": metrics["volumes"],
                "cubagem": metrics["cubagem"],
                "cancelado": metrics["cancelado"],
                "unidade": _clean(cell("unidade")),
                "liquidacao": _clean(cell("liquidacao")),
            }
            _enrich_row_times(rec, row, colmap)
            rec["_cancelado"] = _is_cancelado(rec)
            if _is_junk_row(rec, require_login=True):
                continue
            out.append(rec)
        return out
    finally:
        wb.close()


def _parse_text_455(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise RuntimeError("455: encoding inválido")
    # detect sep
    sample = text.splitlines()[:5]
    sep = ";" if sum(l.count(";") for l in sample) >= sum(l.count(",") for l in sample) else ","
    reader = csv.reader(text.splitlines(), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows[:12]):
        joined = " ".join(_norm_header(c) for c in row if c is not None)
        if (
            "hora de emissao" in joined
            or "serie numero" in joined
            or joined.startswith("ctrc ")
            or joined == "ctrc"
            or " valfrete" in f" {joined}"
            or joined.startswith("ctrc;")
        ):
            header_idx = i
            break
        # slim: primeira linha já é CTRC;EMISS;...
        if i == 0 and any(_norm_header(c) in {"ctrc", "emiss", "valfrete"} for c in row):
            header_idx = 0
            break

    header = rows[header_idx]
    slim = _is_slim_headers(header)
    colmap = _map_headers(header, apply_fixed=not slim)
    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue

        def cell(key: str) -> str:
            idx = colmap.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        ctrc = _clean(cell("ctrc"))
        if not ctrc:
            # SSW text: coluna B = Serie/Numero CTRC (índice 1) no Excel largo
            if not slim and len(row) > 1:
                ctrc = _clean(row[1])
            elif slim and row:
                ctrc = _clean(row[0])
        if not ctrc:
            continue
        up = ctrc.upper()
        if "CTRC" in up and len(ctrc) < 8:
            continue
        login = _login_from_row(row, colmap, slim=slim)
        metrics = _metrics_from_row(row, colmap, slim=slim)
        rec = {
            "ctrc": ctrc,
            "data_emissao": _clean(cell("data_emissao")),
            "hora_emissao": "",
            "data_autorizacao": "",
            "hora_autorizacao": "",
            "expedidor": login,
            "frete": metrics["frete"],
            "valor_mercadoria": metrics["valor_mercadoria"],
            "peso": metrics["peso"],
            "volumes": metrics["volumes"],
            "cubagem": metrics["cubagem"],
            "cancelado": metrics["cancelado"],
            "unidade": _clean(cell("unidade")),
            "liquidacao": _clean(cell("liquidacao")),
        }
        _enrich_row_times(rec, row, colmap)
        rec["_cancelado"] = _is_cancelado(rec)
        # slim costuma não ter login — conta o CTRC mesmo assim
        if _is_junk_row(rec, require_login=not slim):
            continue
        out.append(rec)
    return out


def _parse_br_date(value: Any) -> datetime | None:
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:10], fmt)
        except Exception:
            continue
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%d%m%Y")
        except Exception:
            pass
    if len(digits) == 6:
        try:
            return datetime.strptime(digits, "%d%m%y")
        except Exception:
            pass
    return None


def _period_bounds(periodo: str) -> tuple[datetime | None, datetime | None]:
    """Interpreta periodo_fmt do download ('19/08', '1908', '190826-190826')."""
    raw = (periodo or "").strip()
    if not raw:
        return None, None
    parts = re.split(r"\s*[–\-aA]\s*", raw)
    parts = [p.strip() for p in parts if p.strip()]
    if not parts:
        return None, None

    def one(p: str) -> datetime | None:
        d = _parse_br_date(p)
        if d:
            return d
        dig = re.sub(r"\D", "", p)
        now = datetime.now()
        if len(dig) == 4:  # DDMM
            try:
                return datetime.strptime(dig + now.strftime("%Y"), "%d%m%Y")
            except Exception:
                return None
        if len(dig) == 6:
            try:
                return datetime.strptime(dig, "%d%m%y")
            except Exception:
                return None
        if len(dig) == 8:
            try:
                return datetime.strptime(dig, "%d%m%Y")
            except Exception:
                return None
        return None

    d0 = one(parts[0])
    d1 = one(parts[-1]) if len(parts) > 1 else d0
    return d0, d1


def _row_in_period(rec: dict[str, Any], d0: datetime | None, d1: datetime | None) -> bool:
    if d0 is None and d1 is None:
        return True
    dt = _parse_br_date(rec.get("data_emissao"))
    if dt is None:
        # Com período definido, linha sem data de emissão não entra (evita lixo do CSV slim)
        return False
    day = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    if d0 is not None and day < d0.replace(hour=0, minute=0, second=0, microsecond=0):
        return False
    if d1 is not None and day > d1.replace(hour=0, minute=0, second=0, microsecond=0):
        return False
    return True


def analyze_reports_455(
    files: list[str] | list[Path] | str | Path | None,
    *,
    periodo: str = "",
    on_status=None,
) -> dict[str, Any]:
    status = on_status or (lambda _m: None)
    paths: list[Path] = []
    if isinstance(files, (str, Path)):
        paths = [Path(files)]
    elif files:
        paths = [Path(f) for f in files]

    rows: list[dict[str, Any]] = []
    for p in paths:
        status(f"[455] parse {p.name}")
        rows.extend(parse_excel_455(p))

    d0, d1 = _period_bounds(periodo)
    if d0 or d1:
        before = len(rows)
        rows = [r for r in rows if _row_in_period(r, d0, d1)]
        if before and len(rows) < before:
            status(f"[455] filtro período {periodo}: {before} → {len(rows)} CTRCs")

    # KPIs
    ctes = len(rows)
    peso = sum(float(r.get("peso") or 0) for r in rows)
    valor = sum(float(r.get("valor_mercadoria") or 0) for r in rows)
    volumes = sum(float(r.get("volumes") or 0) for r in rows)
    cubagem = sum(float(r.get("cubagem") or 0) for r in rows)
    frete = sum(float(r.get("frete") or 0) for r in rows)
    cancelados = sum(1 for r in rows if r.get("_cancelado"))
    pendentes = sum(
        1 for r in rows if r.get("_tem_emissao") and not r.get("_tem_autorizacao") and not r.get("_cancelado")
    )
    finalizados = sum(1 for r in rows if r.get("_tem_autorizacao") and not r.get("_cancelado"))

    dia = 0
    noite = 0
    horas = Counter({h: 0 for h in range(24)})
    for r in rows:
        h = r.get("_hora")
        if h is None:
            continue
        horas[int(h)] += 1
        if DIA_START <= int(h) < DIA_END:
            dia += 1
        else:
            noite += 1

    # Expedidores: só logins oficiais · chave = login normalizado
    exp_count: Counter[str] = Counter()
    for r in rows:
        login = _norm_login(str(r.get("expedidor") or ""))
        if _is_expedidor(login):
            exp_count[login] += 1

    total_exp = sum(exp_count.values()) or 1
    expedidores = []
    for login, qtd in exp_count.most_common(20):
        expedidores.append(
            {
                "nome": login,
                "nome_exibicao": _display_expedidor(login),
                "qtd": qtd,
                "pct": round(100.0 * qtd / total_exp, 1),
            }
        )

    horas_rows = [
        {"hora": h, "label": f"{h}:00", "qtd": int(horas.get(h, 0))} for h in range(24)
    ]

    mes_nome = ""
    try:
        mes_nome = MESES_PT[datetime.now().month - 1]
    except Exception:
        mes_nome = ""

    atualizado = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    # periodo SSW "1708" → exibição "17/08"
    periodo_show = (periodo or "").strip()
    if re.fullmatch(r"\d{4}", periodo_show):
        periodo_show = f"{periodo_show[:2]}/{periodo_show[2:]}"
    elif re.fullmatch(r"\d{6}", periodo_show):
        periodo_show = f"{periodo_show[:2]}/{periodo_show[2:4]}/{periodo_show[4:]}"
    resumo = {
        "periodo": periodo_show or periodo or "",
        "mes": mes_nome,
        "atualizado": atualizado,
        "ctes": ctes,
        "peso": round(peso, 2),
        "peso_fmt": _fmt_dec(peso, 2),
        "valor_mercadoria": round(valor, 2),
        "valor_mercadoria_fmt": _fmt_money(valor),
        "volumes": int(round(volumes)),
        "cubagem": round(cubagem, 2),
        "cubagem_fmt": f"{_fmt_dec(cubagem, 2)} m3",
        "frete": round(frete, 2),
        "frete_fmt": _fmt_money(frete),
        "dia": dia,
        "noite": noite,
        "cancelados": cancelados,
        "pendentes": pendentes,
        "finalizados": finalizados,
    }

    # CSV detalhe (sem campos internos)
    detail_rows = []
    for r in rows:
        detail_rows.append(
            {
                "ctrc": r.get("ctrc"),
                "data_emissao": r.get("data_emissao"),
                "hora_emissao": r.get("hora_emissao"),
                "data_autorizacao": r.get("data_autorizacao"),
                "hora_autorizacao": r.get("hora_autorizacao"),
                "expedidor": r.get("expedidor"),
                "frete": r.get("frete"),
                "valor_mercadoria": r.get("valor_mercadoria"),
                "peso": r.get("peso"),
                "volumes": r.get("volumes"),
                "cubagem": r.get("cubagem"),
                "cancelado": "S" if r.get("_cancelado") else "N",
                "unidade": r.get("unidade"),
                "liquidacao": r.get("liquidacao"),
            }
        )

    _write_csv(EMISSOES_455_CSV, EMISSAO_FIELDS, detail_rows)
    _write_csv(RESUMO_455_CSV, RESUMO_FIELDS, [resumo])
    _write_csv(EXPEDIDORES_455_CSV, EXPEDIDOR_FIELDS, expedidores)
    _write_csv(HORAS_455_CSV, HORA_FIELDS, horas_rows)
    LAST_455_JSON.write_text(
        json.dumps(
            {
                "resumo": resumo,
                "expedidores": expedidores,
                "horas": horas_rows,
                "total": ctes,
                "files": [str(p) for p in paths],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    status(
        f"[455] OK · CTEs={ctes} frete={resumo['frete_fmt']} "
        f"dia={dia} noite={noite} cancel={cancelados} "
        f"pend={pendentes} fin={finalizados} exp={len(expedidores)}"
    )
    return {
        "ok": True,
        "total": ctes,
        "resumo": resumo,
        "expedidores": expedidores,
        "horas": horas_rows,
        "rows": detail_rows,
        "files": {
            "emissoes": str(EMISSOES_455_CSV),
            "resumo": str(RESUMO_455_CSV),
            "expedidores": str(EXPEDIDORES_455_CSV),
            "horas": str(HORAS_455_CSV),
        },
    }
