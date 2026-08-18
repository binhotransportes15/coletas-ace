"""Parser Excel SSW 031 — CTRCs com determinada ocorrência."""
from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from config import BASE_DIR, ensure_dirs
from ocorrencias_pendencia import (
    CODIGO_SLA_POSITIVO,
    OCORR_PENDENCIA,
    label_ocorrencia,
    match_codigo_from_text,
    polaridade,
)

CACHE_DIR = BASE_DIR / "data" / "cache"
PENDENCIAS_31_CSV = CACHE_DIR / "pendencias_31.csv"
RESUMO_31_CSV = CACHE_DIR / "resumo_31.csv"
OFENSORES_31_CSV = CACHE_DIR / "ofensores_31.csv"
LAST_31_JSON = CACHE_DIR / "last_run_31.json"

# Colunas Excel (1-based) — fallback posicional
COL_A_CTRC = 1
COL_B_EMISSAO = 2
COL_R_ULTIMA = 18
COL_S_COMPL_ULTIMA = 19
COL_AM_DESC_OCORR = 39
COL_AN_COMPL_OCORR = 40

# Cabeçalhos possíveis do Excel 031 (normalizados)
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "ctrc": ("ctrc", "conhec", "conhecimento", "numero ctrc", "nº ctrc", "n ctrc"),
    "data_emissao": ("emissao", "emissão", "data emissao", "data emissão", "dt emissao"),
    "ultima_ocorrencia": ("ultima ocorr", "última ocorr", "ult ocorr", "ocorrencia atual", "última ocorrência"),
    "historico": ("historico", "histórico", "compl ultima", "compl. ultima", "complemento ultima"),
    "descricao_ocorrencia": ("desc ocorr", "descricao ocorr", "descrição ocorr", "descricao da ocorrencia"),
    "complemento_ocorrencia": ("complemento ocorr", "compl ocorr", "complemento da ocorrencia", "complemento"),
    "nf": ("nro nf", "nr nf", "nota fiscal", "nfe", "nf ", "numero nf", "nº nf"),
    "remetente": ("remetente", "cliente remetente", "embarcador", "expedidor"),
    "destinatario": ("destinatario", "destinatário", "cliente destinatario", "recebedor"),
    "filial": ("filial", "unid", "unidade", "unidade resp", "filial resp", "origem", "unid. resp"),
    "valor_mercadoria": (
        "val merc",
        "valor merc",
        "vl merc",
        "mercadoria",
        "valor da mercadoria",
        "vlr mercadoria",
        "val. mercad",
    ),
    "data_ocorrencia": (
        "data ocorr",
        "dt ocorr",
        "data da ocorrencia",
        "data ocorrência",
        "dt. ocorrencia",
    ),
    "cidade": ("cidade", "municipio", "município"),
    "uf": ("uf", "estado"),
}

PENDENCIA_FIELDS = [
    "ctrc",
    "data_emissao",
    "ultima_ocorrencia",
    "historico",
    "codigo",
    "codigo_consulta",
    "descricao_ocorrencia",
    "complemento_ocorrencia",
    "descricao_codigo",
    "nf",
    "remetente",
    "destinatario",
    "filial",
    "valor_mercadoria",
    "data_ocorrencia",
    "cidade",
    "uf",
    "aging_dias",
    "status_ace",
]

RESUMO_FIELDS = [
    "periodo",
    "atualizado",
    "total_ctrcs",
    "total_codigos",
    "solucionadas",
    "abertas",
    "sla_pct",
    "sla_medio_dias",
    "valor_risco",
    "aging_0_2",
    "aging_3_5",
    "aging_6_mais",
    "topo_codigo",
    "topo_label",
    "topo_qtd",
]

OFENSOR_FIELDS = ["codigo", "label", "qtd", "pct", "polaridade"]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            pass
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _norm_header(text: str) -> str:
    t = str(text or "").strip().lower()
    t = (
        t.replace("á", "a")
        .replace("à", "a")
        .replace("ã", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("õ", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    return re.sub(r"\s+", " ", t)


def _cell(row: tuple[Any, ...] | list[Any], idx: int) -> str:
    if idx < 1 or idx > len(row):
        return ""
    return _clean(row[idx - 1])


def _parse_br_date(raw: str) -> date | None:
    s = str(raw or "").strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if not m:
        return None
    dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if yy < 100:
        yy += 2000
    try:
        return date(yy, mm, dd)
    except ValueError:
        return None


def _parse_money(raw: str) -> float:
    s = str(raw or "").strip()
    if not s:
        return 0.0
    s = re.sub(r"[R$\s]", "", s, flags=re.I)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fmt_money(v: float) -> str:
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _filial_from_ctrc(ctrc: str) -> str:
    m = re.match(r"^([A-Z]{2,4})", str(ctrc or "").upper().replace(" ", ""))
    return m.group(1) if m else ""


def _nf_from_text(*parts: str) -> str:
    blob = " ".join(parts)
    m = re.search(r"\bNFD?\s*[:#-]?\s*(\d{4,})", blob, flags=re.I)
    return m.group(1) if m else ""


def _data_ocorr_from_text(*parts: str) -> str:
    blob = " ".join(parts)
    m = re.search(r"\bEM\s+(\d{1,2}/\d{1,2}/\d{2,4})", blob, flags=re.I)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", blob)
    return m2.group(1) if m2 else ""


def _map_headers(row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    """Retorna field -> índice 1-based se a linha parecer cabeçalho."""
    cells = [_norm_header(_clean(c)) for c in row]
    joined = " | ".join(cells)
    if "ctrc" not in joined and "conhec" not in joined:
        return {}
    out: dict[str, int] = {}
    for i, cell in enumerate(cells, start=1):
        if not cell:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in out:
                continue
            for al in aliases:
                if al in cell or cell in al:
                    out[field] = i
                    break
    return out


def _write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    ensure_dirs()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def _enrich_row(r: dict[str, str], *, codigo_consulta: str = "") -> dict[str, str]:
    ctrc = str(r.get("ctrc") or "").upper().replace(" ", "")
    ultima = r.get("ultima_ocorrencia") or ""
    historico = r.get("historico") or ""
    desc = r.get("descricao_ocorrencia") or ""
    compl = r.get("complemento_ocorrencia") or ""
    codigo = (
        match_codigo_from_text(ultima)
        or match_codigo_from_text(desc)
        or str(r.get("codigo") or codigo_consulta or "").strip()
    )
    nf = r.get("nf") or _nf_from_text(compl, historico, desc)
    filial = r.get("filial") or _filial_from_ctrc(ctrc)
    data_ocorr = r.get("data_ocorrencia") or _data_ocorr_from_text(historico, ultima, compl)
    if not data_ocorr:
        data_ocorr = r.get("data_emissao") or ""
    ref = date.today()
    d0 = _parse_br_date(data_ocorr) or _parse_br_date(r.get("data_emissao") or "")
    aging = (ref - d0).days if d0 else 0
    if aging < 0:
        aging = 0
    status = "solucionada" if codigo == CODIGO_SLA_POSITIVO else "aberta"
    return {
        **r,
        "ctrc": ctrc,
        "codigo": codigo,
        "codigo_consulta": str(codigo_consulta or r.get("codigo_consulta") or "").strip(),
        "descricao_codigo": label_ocorrencia(codigo),
        "nf": nf,
        "filial": filial.upper() if filial else "",
        "data_ocorrencia": data_ocorr,
        "aging_dias": str(aging),
        "status_ace": status,
        "valor_mercadoria": r.get("valor_mercadoria") or "",
        "remetente": r.get("remetente") or "",
        "destinatario": r.get("destinatario") or "",
        "cidade": r.get("cidade") or "",
        "uf": r.get("uf") or "",
    }


def parse_excel_31(path: Path | str, *, codigo_consulta: str = "") -> list[dict[str, str]]:
    """Lê Excel 031: cabeçalho dinâmico + fallback A,B,R,S,AM,AN."""
    from openpyxl import load_workbook

    p = Path(path)
    if not p.is_file():
        raise RuntimeError(f"31: arquivo inexistente: {p}")
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_out: list[dict[str, str]] = []
        header_map: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            if not header_map and i <= 5:
                guessed = _map_headers(row)
                if guessed.get("ctrc"):
                    header_map = guessed
                    continue

            def g(field: str, fallback_idx: int = 0) -> str:
                if header_map.get(field):
                    return _cell(row, header_map[field])
                return _cell(row, fallback_idx) if fallback_idx else ""

            ctrc = g("ctrc", COL_A_CTRC)
            if not ctrc:
                continue
            up = ctrc.upper()
            if i <= 3 and ("CTRC" in up or "CONHEC" in up or "NUMERO" in up or "NÚMERO" in up):
                continue
            base = {
                "ctrc": ctrc,
                "data_emissao": g("data_emissao", COL_B_EMISSAO),
                "ultima_ocorrencia": g("ultima_ocorrencia", COL_R_ULTIMA),
                "historico": g("historico", COL_S_COMPL_ULTIMA),
                "descricao_ocorrencia": g("descricao_ocorrencia", COL_AM_DESC_OCORR),
                "complemento_ocorrencia": g("complemento_ocorrencia", COL_AN_COMPL_OCORR),
                "nf": g("nf"),
                "remetente": g("remetente"),
                "destinatario": g("destinatario"),
                "filial": g("filial"),
                "valor_mercadoria": g("valor_mercadoria"),
                "data_ocorrencia": g("data_ocorrencia"),
                "cidade": g("cidade"),
                "uf": g("uf"),
                "codigo": "",
                "codigo_consulta": str(codigo_consulta or "").strip(),
            }
            rows_out.append(_enrich_row(base, codigo_consulta=codigo_consulta))
        return rows_out
    finally:
        wb.close()


def _parse_any(path: Path, codigo: str) -> list[dict[str, str]]:
    suf = path.suffix.lower()
    if suf in {".xlsx", ".xlsm", ".xls"}:
        return parse_excel_31(path, codigo_consulta=codigo)
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.reader(fh, delimiter=delim)
        out: list[dict[str, str]] = []
        header_map: dict[str, int] = {}
        for i, cols in enumerate(reader, start=1):
            if not cols:
                continue
            if not header_map and i <= 3:
                guessed = _map_headers(cols)
                if guessed.get("ctrc"):
                    header_map = guessed
                    continue

            def g(field: str, fallback_idx: int = 0) -> str:
                if header_map.get(field):
                    idx = header_map[field]
                    return _clean(cols[idx - 1]) if len(cols) >= idx else ""
                return _clean(cols[fallback_idx - 1]) if fallback_idx and len(cols) >= fallback_idx else ""

            ctrc = g("ctrc", COL_A_CTRC)
            if not ctrc or (i <= 2 and "CTRC" in ctrc.upper()):
                continue
            base = {
                "ctrc": ctrc,
                "data_emissao": g("data_emissao", COL_B_EMISSAO),
                "ultima_ocorrencia": g("ultima_ocorrencia", COL_R_ULTIMA),
                "historico": g("historico", COL_S_COMPL_ULTIMA),
                "descricao_ocorrencia": g("descricao_ocorrencia", COL_AM_DESC_OCORR),
                "complemento_ocorrencia": g("complemento_ocorrencia", COL_AN_COMPL_OCORR),
                "nf": g("nf"),
                "remetente": g("remetente"),
                "destinatario": g("destinatario"),
                "filial": g("filial"),
                "valor_mercadoria": g("valor_mercadoria"),
                "data_ocorrencia": g("data_ocorrencia"),
                "cidade": g("cidade"),
                "uf": g("uf"),
                "codigo": "",
                "codigo_consulta": codigo,
            }
            out.append(_enrich_row(base, codigo_consulta=codigo))
        return out


def _dedupe(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by: dict[str, dict[str, str]] = {}
    for r in rows:
        key = r.get("ctrc") or ""
        if not key:
            continue
        prev = by.get(key)
        if not prev:
            by[key] = r
            continue
        score_new = 2 if r.get("codigo") in OCORR_PENDENCIA else 1
        score_old = 2 if prev.get("codigo") in OCORR_PENDENCIA else 1
        if score_new >= score_old:
            by[key] = r
    return list(by.values())


def _format_periodo_display(periodo: str) -> str:
    raw = str(periodo or "").strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 12:
        a, b = digits[:6], digits[6:]
        return f"{a[:2]}/{a[2:4]}/{a[4:]} – {b[:2]}/{b[2:4]}/{b[4:]}"
    if len(digits) == 8:
        a, b = digits[:4], digits[4:]
        return f"{a[:2]}/{a[2:]} – {b[:2]}/{b[2:]}"
    if "-" in raw and "/" not in raw:
        left, _, right = raw.partition("-")
        return f"{left.strip()} – {right.strip()}".strip(" –")
    return raw


def analyze_reports_31(
    paths_by_code: dict[str, str | Path],
    *,
    periodo: str = "",
    on_status: Any = None,
) -> dict[str, Any]:
    status = on_status or (lambda _m: None)
    ensure_dirs()
    all_rows: list[dict[str, str]] = []
    periodo_fmt = _format_periodo_display(periodo)
    for code, path in (paths_by_code or {}).items():
        p = Path(path)
        if not p.is_file():
            status(f"[31/{code}] arquivo ausente — pulou")
            continue
        try:
            chunk = _parse_any(p, str(code))
            status(f"[31/{code}] {len(chunk)} linha(s) em {p.name}")
            all_rows.extend(chunk)
        except Exception as err:  # noqa: BLE001
            status(f"[31/{code}] parse falhou: {err}")

    uniq = [_enrich_row(r) for r in _dedupe(all_rows)]
    for r in uniq:
        if not r.get("codigo"):
            r["codigo"] = match_codigo_from_text(r.get("ultima_ocorrencia") or "") or r.get(
                "codigo_consulta", ""
            )
        r["descricao_codigo"] = label_ocorrencia(r.get("codigo") or "")
        r["status_ace"] = "solucionada" if r.get("codigo") == CODIGO_SLA_POSITIVO else "aberta"

    counts = Counter(str(r.get("codigo") or "").strip() for r in uniq if r.get("codigo"))
    total = len(uniq)
    solucionadas = int(counts.get(CODIGO_SLA_POSITIVO, 0))
    abertas_rows = [r for r in uniq if r.get("status_ace") != "solucionada"]
    abertas = len(abertas_rows)
    sla_pct_num = (100.0 * solucionadas / total) if total else 0.0
    sla_pct = f"{sla_pct_num:.1f}".replace(".", ",")

    valor_risco = sum(_parse_money(r.get("valor_mercadoria") or "") for r in abertas_rows)
    aging_vals = [int(r.get("aging_dias") or 0) for r in abertas_rows]
    sla_medio = (sum(aging_vals) / len(aging_vals)) if aging_vals else 0.0
    aging_0_2 = sum(1 for d in aging_vals if d <= 2)
    aging_3_5 = sum(1 for d in aging_vals if 3 <= d <= 5)
    aging_6 = sum(1 for d in aging_vals if d > 5)

    ofensores: list[dict[str, Any]] = []
    for code, qtd in counts.most_common():
        lab = label_ocorrencia(code) if code else "OUTROS"
        pol = polaridade(code)
        ofensores.append(
            {
                "codigo": code or "?",
                "label": lab,
                "qtd": qtd,
                "pct": f"{(100.0 * qtd / total):.1f}".replace(".", ",") if total else "0,0",
                "polaridade": pol,
            }
        )
    ofensores.sort(
        key=lambda o: (
            0 if o.get("polaridade") == "pos" else 1,
            -(int(o.get("qtd") or 0)),
        )
    )

    ofens_neg = [o for o in ofensores if o.get("polaridade") != "pos"]
    topo = ofens_neg[0] if ofens_neg else {"codigo": "", "label": "—", "qtd": 0}
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    resumo = [
        {
            "periodo": periodo_fmt or "",
            "atualizado": now,
            "total_ctrcs": total,
            "total_codigos": len(ofensores),
            "solucionadas": solucionadas,
            "abertas": abertas,
            "sla_pct": sla_pct,
            "sla_medio_dias": f"{sla_medio:.1f}".replace(".", ","),
            "valor_risco": _fmt_money(valor_risco),
            "aging_0_2": aging_0_2,
            "aging_3_5": aging_3_5,
            "aging_6_mais": aging_6,
            "topo_codigo": topo.get("codigo") or "",
            "topo_label": topo.get("label") or "—",
            "topo_qtd": int(topo.get("qtd") or 0),
        }
    ]

    _write_csv(PENDENCIAS_31_CSV, PENDENCIA_FIELDS, uniq)
    _write_csv(RESUMO_31_CSV, RESUMO_FIELDS, resumo)
    _write_csv(OFENSORES_31_CSV, OFENSOR_FIELDS, ofensores)
    meta = {
        "ok": True,
        "total": total,
        "solucionadas": solucionadas,
        "abertas": abertas,
        "sla_pct": sla_pct,
        "sla_medio_dias": resumo[0]["sla_medio_dias"],
        "valor_risco": resumo[0]["valor_risco"],
        "ofensores": ofensores[:12],
        "resumo": resumo[0],
        "periodo": periodo_fmt,
        "files": {
            "pendencias": str(PENDENCIAS_31_CSV),
            "resumo": str(RESUMO_31_CSV),
            "ofensores": str(OFENSORES_31_CSV),
        },
    }
    LAST_31_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    status(
        f"31 análise: {total} CTRC(s) · SLA {sla_pct}% "
        f"(+{solucionadas} / −{abertas}) · ofensor={topo.get('codigo')}"
    )
    return meta
