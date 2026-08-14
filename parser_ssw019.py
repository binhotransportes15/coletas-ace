"""Parser Excel SSW 019 — CTRCs disponíveis (Sem transferência).

Layout real (sswweb / Excel):
  A=DESTINO · B=CTRC · E=PREV DE ENTREGA · L=DESTINATARIO · M=CIDADE · N=UF · X=OCORRENCIA
Torres: agrega por CIDADE/UF (não pela sigla).
Frete/peso: descobertos pelo cabeçalho.
"""
from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from config import BASE_DIR, ensure_dirs
from siglas_filiais import label_origem, normalizar_sigla

CACHE_DIR = BASE_DIR / "data" / "cache"
ITENS_019_CSV = CACHE_DIR / "itens_019.csv"
RESUMO_019_CSV = CACHE_DIR / "resumo_019.csv"
TOP_CTE_019_CSV = CACHE_DIR / "top_cte_019.csv"
POR_FILIAL_019_CSV = CACHE_DIR / "por_filial_019.csv"
LAST_019_JSON = CACHE_DIR / "last_run_019.json"

COL_FILIAL = "A"
COL_CTRC = "B"
COL_PREVISTA = "E"
COL_CLIENTE = "L"
COL_CIDADE = "M"
COL_UF = "N"
COL_OCORRENCIA = "X"

ITEM_FIELDS = [
    "ctrc",
    "filial",
    "cidade",
    "uf",
    "cidade_uf",
    "cliente",
    "data_prevista",
    "dias_atraso",
    "ultima_ocorrencia",
    "frete",
    "peso",
]
RESUMO_FIELDS = [
    "atualizado",
    "periodo",
    "qtd",
    "frete",
    "frete_fmt",
    "peso",
    "peso_fmt",
]
TOP_CTE_FIELDS = [
    "ctrc",
    "filial",
    "cidade",
    "uf",
    "cidade_uf",
    "cliente",
    "dias_atraso",
    "ultima_ocorrencia",
]
POR_FILIAL_FIELDS = ["cidade", "uf", "cidade_uf", "filial", "qtd", "frete", "peso", "dias_max"]

TOP_N = 10
TOP_CIDADES = 12


def _col_letter_to_idx(letters: str) -> int:
    n = 0
    for ch in str(letters or "").strip().upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - ord("A") + 1)
    return max(0, n - 1)


def _cell_by_letter(row: tuple[Any, ...] | list[Any], letters: str) -> Any:
    idx = _col_letter_to_idx(letters)
    if idx >= len(row):
        return ""
    return row[idx]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            pass
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    try:
        n = float(text or 0)
        return -n if neg else n
    except Exception:
        return 0.0


def _fmt_int(n: float | int) -> str:
    return f"{int(round(float(n) or 0)):,}".replace(",", ".")


def _fmt_dec(n: float, places: int = 2) -> str:
    s = f"{float(n or 0):,.{places}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_money(n: float) -> str:
    return f"R$ {_fmt_dec(n, 2)}"


def _norm_header(h: Any) -> str:
    t = _clean(h).lower()
    for a, b in (
        ("ç", "c"),
        ("ã", "a"),
        ("á", "a"),
        ("à", "a"),
        ("â", "a"),
        ("é", "e"),
        ("ê", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ô", "o"),
        ("ú", "u"),
    ):
        t = t.replace(a, b)
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def _find_cols(headers: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {
        "filial": _col_letter_to_idx(COL_FILIAL),
        "ctrc": _col_letter_to_idx(COL_CTRC),
        "prevista": _col_letter_to_idx(COL_PREVISTA),
        "cliente": _col_letter_to_idx(COL_CLIENTE),
        "cidade": _col_letter_to_idx(COL_CIDADE),
        "uf": _col_letter_to_idx(COL_UF),
        "ocorrencia": _col_letter_to_idx(COL_OCORRENCIA),
    }
    prev_from_entreg = False
    for i, h in enumerate(headers):
        n = _norm_header(h)
        if not n:
            continue
        if n in {"destino", "filial", "unidade", "unid"} or n.startswith("destino"):
            out["filial"] = i
        elif "ctrc" in n or n.startswith("ctrc") or "gai" in n:
            out["ctrc"] = i
        elif "prev" in n and "entreg" in n:
            out["prevista"] = i
            prev_from_entreg = True
        elif (not prev_from_entreg) and "prev" in n and ("cheg" in n or n.startswith("prev")):
            out["prevista"] = i
        elif n in {"destinatario", "cliente", "recebedor"} or "destinat" in n:
            out["cliente"] = i
        elif n == "cidade" or n.startswith("cidade"):
            out["cidade"] = i
        elif n in {"uf", "estado"}:
            out["uf"] = i
        elif "ocorr" in n:
            out["ocorrencia"] = i
        elif "frete" in n or n in {"vlr frete", "valor frete", "vl frete"}:
            out.setdefault("frete", i)
        elif n == "peso" or n.startswith("peso ") or "peso real" in n or "peso taxado" in n:
            out.setdefault("peso", i)
    return out


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, "date") and callable(getattr(value, "date")):
        try:
            d = value.date()
            if isinstance(d, date):
                return d
        except Exception:
            pass
    text = _clean(value)
    if not text:
        return None
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if m:
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return date(yy, mm, dd)
        except Exception:
            return None
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        try:
            return date(int(digits[4:8]), int(digits[2:4]), int(digits[0:2]))
        except Exception:
            return None
    if len(digits) == 6:
        try:
            return date(2000 + int(digits[4:6]), int(digits[2:4]), int(digits[0:2]))
        except Exception:
            return None
    return None


def _dias_atraso(prevista: date | None, hoje: date | None = None) -> int:
    if prevista is None:
        return 0
    ref = hoje or date.today()
    return max(0, int((ref - prevista).days))


def _looks_ctrc(value: str) -> bool:
    t = str(value or "").strip().upper()
    if not t or t in {"N", "S", "X", "T", "CTRC", "CTRC/GAI/PAL"}:
        return False
    if re.match(r"^[A-Z]{2,4}\d{4,}(-\d)?$", t):
        return True
    if re.search(r"\d{5,}", t) and len(t) >= 6:
        return True
    return False


def _cidade_uf(cidade: str, uf: str, filial: str = "") -> str:
    c = _clean(cidade).upper()
    u = _clean(uf).upper()
    if c and u:
        return f"{c}/{u}"
    if c:
        return c
    if u:
        return u
    lab = label_origem(filial)
    if lab and u:
        return f"{lab}/{u}"
    return lab or filial or "—"


def _write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    ensure_dirs()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def parse_excel_019(path: Path | str) -> list[dict[str, Any]]:
    p = Path(path)
    if not p.is_file():
        return []
    suf = p.suffix.lower()
    if suf in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx_019(p)
    return _parse_text_019(p)


def _row_to_item(cells: list[Any], cols: dict[str, int], hoje: date) -> dict[str, Any] | None:
    def _at(key: str) -> Any:
        idx = cols.get(key, -1)
        if idx is None or idx < 0 or idx >= len(cells):
            return ""
        return cells[idx]

    ctrc = _clean(_at("ctrc"))
    if not _looks_ctrc(ctrc):
        return None
    filial = normalizar_sigla(_clean(_at("filial")))
    if not filial:
        m = re.match(r"^([A-Z]{2,4})\d", ctrc.upper())
        filial = m.group(1) if m else ""
    cidade = _clean(_at("cidade")).upper()
    uf = _clean(_at("uf")).upper()
    if not cidade:
        cidade = (label_origem(filial) or "").upper()
    cidade_uf = _cidade_uf(cidade, uf, filial)
    cliente = _clean(_at("cliente"))
    raw_prev = _at("prevista")
    prevista = _parse_date(raw_prev)
    ocorr = _clean(_at("ocorrencia"))
    frete = _num(_at("frete")) if "frete" in cols else 0.0
    peso = _num(_at("peso")) if "peso" in cols else 0.0
    if isinstance(raw_prev, (date, datetime)) and prevista:
        data_prevista = prevista.strftime("%d/%m/%Y")
    else:
        data_prevista = _clean(raw_prev)
    return {
        "ctrc": ctrc,
        "filial": filial,
        "cidade": cidade,
        "uf": uf,
        "cidade_uf": cidade_uf,
        "cliente": cliente,
        "data_prevista": data_prevista,
        "dias_atraso": _dias_atraso(prevista, hoje),
        "ultima_ocorrencia": ocorr,
        "frete": frete,
        "peso": peso,
    }


def _parse_xlsx_019(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter) or ())
        except StopIteration:
            return []
        cols = _find_cols(headers)
        hoje = date.today()
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if not row:
                continue
            item = _row_to_item(list(row), cols, hoje)
            if item:
                out.append(item)
        return out
    finally:
        wb.close()


def _parse_text_019(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise RuntimeError("019: encoding inválido")
    sample = text.splitlines()[:5]
    sep = ";" if sum(l.count(";") for l in sample) >= sum(l.count(",") for l in sample) else ","
    reader = csv.reader(text.splitlines(), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    cols = _find_cols(headers)
    hoje = date.today()
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row:
            continue
        item = _row_to_item(row, cols, hoje)
        if item:
            out.append(item)
    return out


def analyze_reports_019(
    files: list[str] | list[Path] | str | Path | None,
    *,
    periodo: str = "",
    on_status=None,
) -> dict[str, Any]:
    status = on_status or (lambda _m: None)
    paths: list[Path] = []
    if isinstance(files, (str, Path)):
        paths = [Path(files)]
    elif files:
        paths = [Path(f) for f in files]

    rows: list[dict[str, Any]] = []
    for p in paths:
        if not p.is_file():
            continue
        status(f"[019] parse {p.name}")
        rows.extend(parse_excel_019(p))

    qtd = len(rows)
    frete = sum(float(r.get("frete") or 0) for r in rows)
    peso = sum(float(r.get("peso") or 0) for r in rows)

    top_cte = sorted(
        rows,
        key=lambda r: (int(r.get("dias_atraso") or 0), str(r.get("ctrc") or "")),
        reverse=True,
    )[:TOP_N]
    top_cte_out = [
        {
            "ctrc": r.get("ctrc") or "",
            "filial": r.get("filial") or "",
            "cidade": r.get("cidade") or "",
            "uf": r.get("uf") or "",
            "cidade_uf": r.get("cidade_uf") or "",
            "cliente": r.get("cliente") or "",
            "dias_atraso": int(r.get("dias_atraso") or 0),
            "ultima_ocorrencia": r.get("ultima_ocorrencia") or "",
        }
        for r in top_cte
    ]

    # Torres: CTRCs por CIDADE/UF
    by_cid: dict[str, dict[str, Any]] = {}
    for r in rows:
        key = str(r.get("cidade_uf") or "").strip() or "—"
        slot = by_cid.setdefault(
            key,
            {
                "cidade": r.get("cidade") or "",
                "uf": r.get("uf") or "",
                "cidade_uf": key,
                "filial": r.get("filial") or "",
                "qtd": 0,
                "frete": 0.0,
                "peso": 0.0,
                "dias_max": 0,
            },
        )
        slot["qtd"] += 1
        slot["frete"] += float(r.get("frete") or 0)
        slot["peso"] += float(r.get("peso") or 0)
        slot["dias_max"] = max(int(slot["dias_max"]), int(r.get("dias_atraso") or 0))

    por_filial = sorted(
        by_cid.values(),
        key=lambda x: (int(x["qtd"]), int(x["dias_max"])),
        reverse=True,
    )[:TOP_CIDADES]

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumo = {
        "atualizado": agora,
        "periodo": periodo or "",
        "qtd": qtd,
        "frete": frete,
        "frete_fmt": _fmt_money(frete),
        "peso": peso,
        "peso_fmt": _fmt_int(peso),
    }

    ensure_dirs()
    _write_csv(ITENS_019_CSV, ITEM_FIELDS, rows)
    _write_csv(RESUMO_019_CSV, RESUMO_FIELDS, [resumo])
    _write_csv(TOP_CTE_019_CSV, TOP_CTE_FIELDS, top_cte_out)
    _write_csv(POR_FILIAL_019_CSV, POR_FILIAL_FIELDS, por_filial)

    payload = {
        "ok": True,
        "programa": "019",
        "view": "sem_transferencia",
        "resumo": resumo,
        "top_cte": top_cte_out,
        "por_filial": por_filial,
        "itens": rows,
        "total": qtd,
    }
    LAST_019_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    status(
        f"[019] OK · CTRCs={qtd} cidades={len(by_cid)} "
        f"frete={resumo['frete_fmt']} peso={resumo['peso_fmt']}"
    )
    return payload
